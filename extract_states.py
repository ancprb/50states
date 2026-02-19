import openpyxl
import json

wb = openpyxl.load_workbook('/Users/andrewperabeau/Desktop/50 States Project/50_Economies_State_Data.xlsx')

# ── 1. State Overview ──────────────────────────────────────────────────────────
overview_ws = wb['State Overview']
overview_data = {}
for row in overview_ws.iter_rows(min_row=5, values_only=True):
    if row[1] is None:
        continue
    state_name = row[1]
    pop_change_raw = row[6]
    # Strip "%" and convert to float
    if isinstance(pop_change_raw, str):
        pop_change = float(pop_change_raw.replace('%', ''))
    else:
        pop_change = float(pop_change_raw) if pop_change_raw is not None else None

    overview_data[state_name] = {
        'rank':             row[0],
        'name':             state_name,
        'abbreviation':     row[2],
        'capital':          row[3],
        'region':           row[4],
        'population':       int(row[5]) if row[5] is not None else None,
        'popChange':        pop_change,
        'gdp':              float(row[7]) if row[7] is not None else None,
        'gdpPerCapita':     int(row[8])   if row[8] is not None else None,
        'medianIncome':     int(row[9])   if row[9] is not None else None,
        'unemploymentRate': float(row[10]) if row[10] is not None else None,
        'obesityRate':      float(row[11]) if row[11] is not None else None,
        'uninsuredRate':    float(row[12]) if row[12] is not None else None,
    }

# ── 2. Economic Data ──────────────────────────────────────────────────────────
econ_ws = wb['Economic Data']
econ_data = {}
for row in econ_ws.iter_rows(min_row=5, values_only=True):
    if row[0] is None:
        continue
    econ_data[row[0]] = {
        'topIndustries':    row[5],
        'majorEmployers':   row[6],
        'economicSource':   row[7],
    }

# ── 3. Health Data ────────────────────────────────────────────────────────────
health_ws = wb['Health Data']
health_data = {}
for row in health_ws.iter_rows(min_row=5, values_only=True):
    if row[0] is None:
        continue
    health_data[row[0]] = {
        'lifeExpectancy':       float(row[3]) if row[3] is not None else None,
        'obesitySource':        row[4],
        'uninsuredSource':      row[5],
        'lifeExpectancySource': row[6],
    }

# ── 4. Industry Data ──────────────────────────────────────────────────────────
industry_ws = wb['Industry Data']
industry_data = {}
for row in industry_ws.iter_rows(min_row=5, values_only=True):
    if row[0] is None:
        continue
    industry_data[row[0]] = {
        'industrySource': row[3],
    }

# ── 5. Fun Facts ──────────────────────────────────────────────────────────────
facts_ws = wb['Fun Facts']
facts_data = {}
for row in facts_ws.iter_rows(min_row=5, values_only=True):
    if row[0] is None:
        continue
    facts_data[row[0]] = {
        'funFact1':       row[1],
        'funFact1Source': row[2],
        'funFact2':       row[3],
        'funFact2Source': row[4],
        'funFact3':       row[5],
        'funFact3Source': row[6],
    }

# ── 6. Source Links ───────────────────────────────────────────────────────────
sources_ws = wb['Source Links']
sources = []
src_id = 1
for row in sources_ws.iter_rows(min_row=4, values_only=True):
    if row[0] is None and row[1] is None:
        continue
    if row[0] == 'Data Category':   # skip header row if repeated
        continue
    sources.append({
        'id':       src_id,
        'category': row[0],
        'name':     row[1],
        'url':      row[2],
    })
    src_id += 1

# ── 7. Merge all data into unified states list ────────────────────────────────
states_list = []
for state_name, base in overview_data.items():
    e = econ_data.get(state_name, {})
    h = health_data.get(state_name, {})
    ind = industry_data.get(state_name, {})
    f = facts_data.get(state_name, {})

    entry = {
        'name':               base['name'],
        'abbreviation':       base['abbreviation'],
        'capital':            base['capital'],
        'region':             base['region'],
        'population':         base['population'],
        'popChange':          base['popChange'],
        'gdp':                base['gdp'],
        'gdpPerCapita':       base['gdpPerCapita'],
        'medianIncome':       base['medianIncome'],
        'unemploymentRate':   base['unemploymentRate'],
        'obesityRate':        base['obesityRate'],
        'uninsuredRate':      base['uninsuredRate'],
        'lifeExpectancy':     h.get('lifeExpectancy'),
        'topIndustries':      e.get('topIndustries'),
        'majorEmployers':     e.get('majorEmployers'),
        'funFact1':           f.get('funFact1'),
        'funFact1Source':     f.get('funFact1Source'),
        'funFact2':           f.get('funFact2'),
        'funFact2Source':     f.get('funFact2Source'),
        'funFact3':           f.get('funFact3'),
        'funFact3Source':     f.get('funFact3Source'),
        # Source URLs embedded per state
        'sources': {
            'economic':        e.get('economicSource'),
            'obesity':         h.get('obesitySource'),
            'uninsured':       h.get('uninsuredSource'),
            'lifeExpectancy':  h.get('lifeExpectancySource'),
            'industry':        ind.get('industrySource'),
        }
    }
    states_list.append(entry)

# Sort alphabetically by state name
states_list.sort(key=lambda x: x['name'])

output = {
    'states': states_list,
    'sources': sources
}

out_path = '/Users/andrewperabeau/Desktop/50 States Project/states_data.json'
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(output, f, indent=2, ensure_ascii=False)

print(f"Total states extracted: {len(states_list)}")
print(f"Total sources extracted: {len(sources)}")
print(f"\nOutput written to: {out_path}")
print("\n=== First 2 state entries ===")
print(json.dumps(states_list[:2], indent=2, ensure_ascii=False))
